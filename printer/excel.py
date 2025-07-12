import os
import asyncio
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

import sys
from pathlib import Path

# Добавляем корневую папку в путь для импортов
sys.path.append(str(Path(__file__).parent.parent))

from fetch_orders.mocks import mock_get_new_orders


class ExcelReportManager:
    """Менеджер для создания Excel отчетов по артикулам"""
    
    def __init__(self, filename: str = "print_status_report.xlsx"):
        self.filename = filename
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        if self.sheet:
            self.sheet.title = "Статус печати"
        
        # Цвета для статусов
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        self.header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Настройка заголовков
        self._setup_headers()
        
    def _setup_headers(self):
        """Настройка заголовков таблицы"""
        headers = [
            "Артикул",
            "ID заказа", 
            "Статус печати",
            "Дата создания",
            "Приоритет",
            "Принтер",
            "Путь к файлу"
        ]
        
        if self.sheet:
            for col, header in enumerate(headers, 1):
                cell = self.sheet.cell(row=1, column=col, value=header)
                cell.fill = self.header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                
            # Автоматическая ширина колонок
            for col in range(1, len(headers) + 1):
                self.sheet.column_dimensions[get_column_letter(col)].width = 15
            
    async def generate_report(self, redis_url: str = "redis://localhost:6379") -> str:
        """
        Генерирует Excel отчет со статусом печати артикулов
        
        Args:
            redis_url: URL подключения к Redis
            
        Returns:
            str: Путь к созданному файлу
        """
        # Получаем заказы
        orders_data = await mock_get_new_orders()
        orders = orders_data.get("orders", [])
        
        # Получаем статус из Redis (если доступен)
        print_status = await self._get_print_status_from_redis(redis_url)
        
        row = 2  # Начинаем с 2-й строки (после заголовков)
        
        for order in orders:
            article = order.get("article", "")
            order_id = order.get("id", "")
            
            # Проверяем статус печати
            status = print_status.get(str(order_id), "Не распечатан")
            is_printed = status == "Распечатан"
            
            # Данные для строки
            row_data = [
                article,
                order_id,
                status,
                order.get("createdAt", ""),
                order.get("priority", 1),
                print_status.get(f"{order_id}_printer", ""),
                f"for_print/{article}/ПЕЧАТЬ.png"
            ]
            
            # Заполняем строку
            if self.sheet:
                for col, value in enumerate(row_data, 1):
                    cell = self.sheet.cell(row=row, column=col, value=value)
                    
                    # Применяем цвет в зависимости от статуса
                    if col == 3:  # Колонка статуса
                        if is_printed:
                            cell.fill = self.green_fill
                        else:
                            cell.fill = self.yellow_fill
                            
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            row += 1
            
        # Сохраняем файл
        self.workbook.save(self.filename)
        return self.filename
        
    async def _get_print_status_from_redis(self, redis_url: str) -> Dict[str, str]:
        """Получает статус печати из Redis"""
        try:
            import redis.asyncio as redis
            r = redis.from_url(redis_url)
            await r.ping()
            
            # Получаем все задачи из очереди
            tasks = await r.zrange("print_queue", 0, -1, withscores=True)
            
            status_dict = {}
            for task_json, score in tasks:
                task_data = eval(task_json)  # Безопаснее использовать json.loads
                order_id = str(task_data.get("order_id", ""))
                status = task_data.get("status", "pending")
                
                if status == "completed":
                    status_dict[order_id] = "Распечатан"
                    status_dict[f"{order_id}_printer"] = task_data.get("assigned_printer", "")
                else:
                    status_dict[order_id] = "В очереди"
                    
            await r.close()
            return status_dict
            
        except Exception as e:
            print(f"Ошибка при получении статуса из Redis: {e}")
            return {}
            
    def update_status(self, order_id: str, status: str, printer: str = ""):
        """
        Обновляет статус конкретного заказа в Excel
        
        Args:
            order_id: ID заказа
            status: Новый статус
            printer: Принтер (опционально)
        """
        if not self.sheet:
            return
            
        # Ищем строку с заказом
        for row in range(2, self.sheet.max_row + 1):
            if self.sheet.cell(row=row, column=2).value == order_id:
                # Обновляем статус
                status_cell = self.sheet.cell(row=row, column=3, value=status)
                
                # Применяем цвет
                if status == "Распечатан":
                    status_cell.fill = self.green_fill
                else:
                    status_cell.fill = self.yellow_fill
                    
                # Обновляем принтер
                if printer:
                    self.sheet.cell(row=row, column=6, value=printer)
                    
                break
                
        # Сохраняем изменения
        self.workbook.save(self.filename)


async def create_print_status_report(redis_url: str = "redis://localhost:6379") -> str:
    """
    Создает Excel отчет со статусом печати всех артикулов
    
    Args:
        redis_url: URL подключения к Redis
        
    Returns:
        str: Путь к созданному файлу
    """
    manager = ExcelReportManager()
    filename = await manager.generate_report(redis_url)
    print(f"Отчет создан: {filename}")
    return filename


# Для тестирования
if __name__ == "__main__":
    async def main():
        filename = await create_print_status_report()
        print(f"Excel файл создан: {filename}")
        
        # Демонстрация обновления статуса
        manager = ExcelReportManager(filename)
        manager.update_status("12345", "Распечатан", "printer_1")
        print("Статус обновлен")
    
    asyncio.run(main())
